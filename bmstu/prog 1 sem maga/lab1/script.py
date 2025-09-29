import os
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from bs4 import BeautifulSoup
import openpyxl
from docx import Document

class DocumentVectorizer:
    def __init__(self):
        self.vocabulary = set()
        self.word_to_index = {}
        self.index_to_word = {}
        self.stop_words = {
            'и', 'в', 'во', 'не', 'что', 'он', 'на', 'я', 'с', 'со', 'как', 'а', 
            'то', 'все', 'она', 'так', 'его', 'но', 'да', 'ты', 'к', 'у', 'же', 
            'вы', 'за', 'бы', 'по', 'только', 'ее', 'мне', 'было', 'вот', 'от', 
            'меня', 'еще', 'нет', 'о', 'из', 'ему', 'теперь', 'когда', 'даже', 
            'ну', 'вдруг', 'ли', 'если', 'уже', 'или', 'ни', 'быть', 'был', 
            'него', 'до', 'вас', 'нибудь', 'опять', 'уж', 'вам', 'ведь', 'там', 
            'потом', 'себя', 'ничего', 'ей', 'может', 'они', 'тут', 'где', 'есть', 
            'надо', 'ней', 'для', 'мы', 'тебя', 'их', 'чем', 'была', 'сам', 'чтоб', 
            'без', 'будто', 'чего', 'раз', 'тоже', 'себе', 'под', 'будет', 'ж', 
            'тогда', 'кто', 'этот', 'того', 'потому', 'этого', 'какой', 'совсем', 
            'ним', 'здесь', 'этом', 'один', 'почти', 'мой', 'тем', 'чтобы', 'нее', 
            'сейчас', 'были', 'куда', 'зачем', 'всех', 'никогда', 'можно', 'при', 
            'наконец', 'два', 'об', 'другой', 'хоть', 'после', 'над', 'больше', 
            'тот', 'через', 'эти', 'нас', 'про', 'всего', 'них', 'какая', 'много', 
            'разве', 'три', 'эту', 'моя', 'впрочем', 'хорошо', 'свою', 'этой', 
            'перед', 'иногда', 'лучше', 'чуть', 'том', 'нельзя', 'такой', 'им', 
            'более', 'всегда', 'конечно', 'всю', 'между'
        }
        
    def extract_word_forms(self, text):
        """Извлечение словоформ (токенизация с сохранением исходных форм)"""
        # Приведение к нижнему регистру и разделение на слова
        words = re.findall(r'\b[а-яёa-z]+\b', text.lower())
        # Удаление стоп-слов
        words = [word for word in words if word not in self.stop_words and len(word) > 2]
        return words
    
    def extract_lemma_forms(self, text):
        """Извлечение начальных форм слов (простая лемматизация)"""
        words = self.extract_word_forms(text)
        lemmas = []
        
        for word in words:
            lemma = self.simple_lemmatize(word)
            lemmas.append(lemma)
            
        return lemmas
    
    def simple_lemmatize(self, word):
        """Простая лемматизация для русского языка"""
        if len(word) <= 3:
            return word
            
        # Правила для существительных (женский род)
        if word.endswith('ости') or word.endswith('асти'):
            return word[:-2] + 'ость'
        elif word.endswith('ации'):
            return word[:-3] + 'ация'
        elif word.endswith('ии'):
            return word[:-1]
            
        # Правила для прилагательных
        if word.endswith('ого') or word.endswith('его'):
            return word[:-3] + 'ий'
        elif word.endswith('ым') or word.endswith('им'):
            return word[:-2] + 'ый'
        elif word.endswith('ой') or word.endswith('ей'):
            return word[:-2] + 'ый'
            
        # Правила для глаголов
        if word.endswith('ить') or word.endswith('еть') or word.endswith('ать'):
            return word[:-2] + 'ь'
        elif word.endswith('ют') or word.endswith('ут'):
            return word[:-2]
        elif word.endswith('ят') or word.endswith('ат'):
            return word[:-2]
        elif word.endswith('ил') or word.endswith('ел'):
            return word[:-2] + 'ь'
            
        # Удаление окончаний множественного числа
        if word.endswith('ы') or word.endswith('и'):
            base = word[:-1]
            if len(base) > 2:
                return base
                
        return word
    
    def build_vocabulary(self, documents):
        """Построение словаря из всех документов"""
        all_words = []
        for doc in documents:
            all_words.extend(self.extract_word_forms(doc))
            all_words.extend(self.extract_lemma_forms(doc))
        
        self.vocabulary = set(all_words)
        self.word_to_index = {word: idx for idx, word in enumerate(sorted(self.vocabulary))}
        self.index_to_word = {idx: word for word, idx in self.word_to_index.items()}
    
    def create_word_vectors(self, documents, method='word_forms'):
        """Создание векторов документов"""
        vectors = []
        
        for doc in documents:
            if method == 'word_forms':
                words = self.extract_word_forms(doc)
            else:  # lemma_forms
                words = self.extract_lemma_forms(doc)
            
            # Создание вектора частот слов
            vector = [0] * len(self.vocabulary)
            word_counts = Counter(words)
            
            for word, count in word_counts.items():
                if word in self.word_to_index:
                    vector[self.word_to_index[word]] = count
            
            vectors.append(vector)
        
        return vectors

class DataProcessor:
    def __init__(self):
        self.vectorizer = DocumentVectorizer()
    
    def read_docx_file(self, file_path):
        """Чтение DOCX файла"""
        try:
            doc = Document(file_path)
            text = []
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
            return '\n'.join(text)
        except Exception as e:
            print(f"Ошибка чтения DOCX файла {file_path}: {e}")
            return ""
    
    def read_xml_file(self, file_path):
        """Чтение XML файла с разметкой"""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            return self.parse_xml_content(root)
        except Exception as e:
            print(f"Ошибка чтения XML файла {file_path}: {e}")
            return ""
    
    def parse_xml_content(self, root):
        """Парсинг XML контента"""
        text_parts = []
        
        for elem in root.iter():
            if elem.text and elem.text.strip():
                text_parts.append(elem.text.strip())
            if elem.tail and elem.tail.strip():
                text_parts.append(elem.tail.strip())
        
        return ' '.join(text_parts)
    
    def process_dataset(self, dataset_path):
        """Обработка всего набора данных"""
        documents = []
        document_names = []
        
        if os.path.isfile(dataset_path):
            # Одиночный файл
            if dataset_path.endswith('.docx'):
                text = self.read_docx_file(dataset_path)
                if text:
                    documents.append(text)
                    document_names.append(os.path.basename(dataset_path))
        
        elif os.path.isdir(dataset_path):
            # Папка с файлами
            for file_name in os.listdir(dataset_path):
                file_path = os.path.join(dataset_path, file_name)
                if file_name.endswith('.docx'):
                    text = self.read_docx_file(file_path)
                    if text:
                        documents.append(text)
                        document_names.append(file_name)
        
        return documents, document_names
    
    def create_report(self, documents, document_names, output_path):
        """Создание отчёта в Excel"""
        # Построение словаря
        self.vectorizer.build_vocabulary(documents)
        
        # Создание векторов
        word_vectors = self.vectorizer.create_word_vectors(documents, 'word_forms')
        lemma_vectors = self.vectorizer.create_word_vectors(documents, 'lemma_forms')
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        
        # Лист с словоформами
        ws_word = wb.active
        ws_word.title = "Словоформы"
        
        # Заголовки
        headers = ['Документ'] + [self.vectorizer.index_to_word[i] for i in range(len(self.vectorizer.vocabulary))]
        ws_word.append(headers)
        
        # Данные
        for i, (doc_name, vector) in enumerate(zip(document_names, word_vectors)):
            row = [doc_name] + vector
            ws_word.append(row)
        
        # Лист с начальными формами
        ws_lemma = wb.create_sheet("Начальные_формы")
        ws_lemma.append(headers)
        
        for i, (doc_name, vector) in enumerate(zip(document_names, lemma_vectors)):
            row = [doc_name] + vector
            ws_lemma.append(row)
        
        # Лист со статистикой
        ws_stats = wb.create_sheet("Статистика")
        ws_stats.append(['Параметр', 'Значение'])
        ws_stats.append(['Всего документов', len(documents)])
        ws_stats.append(['Размер словаря', len(self.vectorizer.vocabulary)])
        ws_stats.append(['Общее количество слов (словоформы)', 
                        sum(len(self.vectorizer.extract_word_forms(doc)) for doc in documents)])
        ws_stats.append(['Общее количество слов (леммы)', 
                        sum(len(self.vectorizer.extract_lemma_forms(doc)) for doc in documents)])
        
        # Сохранение
        wb.save(output_path)
        
        return word_vectors, lemma_vectors

def main():
    processor = DataProcessor()
    
    # Обработка примера документа
    '''sample_text = """
    Связь медицины и физической культуры уходит корнями в глубокую древность. 
    Гиппократ, который признан отцом медицины, жил в пятом веке до нашей эры.
    """'''
    sample_text = processor.read_docx_file('C:\Users\MPyankov\Desktop\тексты\Спортивная медицина\спортивная_медицина_05.docx')
    
    print("=== ТЕСТИРОВАНИЕ МЕТОДОВ ===")
    print("Исходный текст:", sample_text)
    print("\nСловоформы:", processor.vectorizer.extract_word_forms(sample_text))
    print("\nНачальные формы:", processor.vectorizer.extract_lemma_forms(sample_text))
    
    # Создание отчёта для примера
    documents = [sample_text]
    document_names = ["пример_документа"]
    
    # Сохранение результатов
    processor.create_report(documents, document_names, "векторы_документов.xlsx")
    print("\nОтчёт сохранён в файл 'векторы_документов.xlsx'")

if __name__ == "__main__":
    main()